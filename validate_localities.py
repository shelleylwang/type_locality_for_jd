#!/usr/bin/env python3
"""
validate_localities.py
=======================
Validates and corrects country assignments for tetrapod type localities
using Claude Opus 4.6 via the Anthropic API.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 STEP 1 — GET AN ANTHROPIC API KEY
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 1. Go to https://console.anthropic.com and create a free account.
 2. In the left sidebar, click "API Keys".
 3. Click "Create Key", give it a name (e.g. "type-locality"), copy it.
    It looks like: sk-ant-api03-...
 4. Add a payment method under "Billing" and deposit a small amount
    (e.g. $20). The script won't exceed your daily budget automatically.
 5. Set the key as an environment variable so this script can use it:
      Mac/Linux terminal:  export ANTHROPIC_API_KEY="sk-ant-api03-..."
      Windows CMD:         setx ANTHROPIC_API_KEY "sk-ant-api03-..."
    Alternatively, paste it directly into the ANTHROPIC_API_KEY line
    in the CONFIGURATION section below. Either way works.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 STEP 2 — INSTALL DEPENDENCIES (one time only)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 In your terminal, run:
   pip install anthropic openpyxl

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 STEP 3 — RUN THE SCRIPT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 Put this script in the same folder as your Excel file, then run:
   python validate_localities.py

 The script processes rows until it hits today's spending limit, then
 stops. Run it again tomorrow — it resumes exactly where it left off.
 When complete, check Tetrapods_validated.xlsx for results.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 WHAT THE SCRIPT PRODUCES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Tetrapods_validated.xlsx  — your original file with two new columns
                               added per sheet: Corrected_Country and Comment
  checkpoint.json           — tracks progress between runs (do not delete)
"""

import os
import re
import sys
import json
import time
import signal
import anthropic
import openpyxl
from datetime import date
from pathlib import Path


# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
#  Edit these values before running.
# ══════════════════════════════════════════════════════════════════════════════

# Your Anthropic API key. Uses the environment variable by default.
# If you want to hard-code it instead, replace "" with your key string.
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")

# File names — keep the script and Excel file in the same folder.
INPUT_FILE      = "Tetrapods_01022026.xlsx"
OUTPUT_FILE     = "Tetrapods_validated.xlsx"
CHECKPOINT_FILE = "checkpoint.json"

# Maximum USD to spend today before the script stops automatically.
# Run the script again tomorrow to continue. It will reset this counter.
# At ~$0.001–0.002 per row with Opus 4.6:
#   $5/day  ≈ 2,500–5,000 rows   (completes in ~8–16 days)
#   $10/day ≈ 5,000–10,000 rows  (completes in ~4–8 days)
#   $20/day ≈ 10,000–20,000 rows (completes in ~2–4 days)
DAILY_BUDGET_USD = 10.00

# How many species rows to pack into a single API call.
# 20 rows is the sweet spot: fast, cost-efficient, and accurate.
# Reduce to 10 if you see quality issues; increase to 30 if you want speed.
ROWS_PER_CALL = 20

# The Claude model to use. Opus 4.6 has the best knowledge of historical
# geography, colonial place names, and species distributions.
MODEL = "claude-opus-4-6"

# Claude Opus 4.6 standard pricing per million tokens.
# ($5.00 input / $25.00 output as of early 2026)
INPUT_PRICE_PER_TOKEN  = 5.00  / 1_000_000
OUTPUT_PRICE_PER_TOKEN = 25.00 / 1_000_000

# NOTE ON FURTHER COST SAVINGS:
# The Anthropic Batch API offers a 50% discount but returns results
# asynchronously (up to 24 hours). This script uses the real-time API
# for simplicity and immediate feedback. If you want to cut costs in
# half and don't mind a more complex workflow, the Batch API is an option.


# ══════════════════════════════════════════════════════════════════════════════
#  THE ANALYSIS PROMPT
#
#  This is the instruction set sent to Claude before every batch of rows.
#  It is the "system prompt" — persistent background context that Claude
#  keeps in mind while analyzing every row. Think of it as briefing an
#  expert consultant before handing them the data.
# ══════════════════════════════════════════════════════════════════════════════

SYSTEM_PROMPT = """You are an expert biogeographer and taxonomist helping to validate type localities for tetrapod species (reptiles, mammals, amphibians, and birds).

Task: You will receive a numbered list of species entries. Each entry shows the species name, year of description, currently assigned country, and the raw type locality text. For every entry, determine whether the assigned country is correct and correct it if needed.

Rules — follow these carefully:
1. Nominal Species: If the locality text mentions multiple subspecies, validate only the locality of the nominal subspecies (the one where the subspecific epithet matches the specific epithet, usually listed first).
2. Language & Typos: Localities may be in any language (Latin, German, French, Spanish, Portuguese, etc.) and may contain typos (e.g., "Columbia" instead of "Colombia"). Interpret the true location from context.
3. Infer from Landmarks: If no country is named explicitly but a city, river, mountain, region, or other landmark is mentioned, infer the correct modern country or territory from that landmark.
4. Historical & Colonial Names: Resolve historical place names into modern equivalents using the year of description as a guide. Examples: "Portuguese Guinea" (1950) → "Guinea-Bissau"; "Siam" → "Thailand"; "Dahomey" → "Benin"; "Rhodesia" → "Zimbabwe"; "British Guiana" → "Guyana"; "Ceylon" → "Sri Lanka".
5. Geopolitical Edge Cases: Handle carefully: Taiwan (use "Taiwan" not "China"), Palestine/West Bank/Gaza, Western Sahara, Kosovo, South Ossetia, Somaliland, Guayana Esequiba, and similar disputed or sensitive territories.
6. Misleading Ports: Some historical localities name the port where the specimen arrived in Europe rather than the actual collection site. Use species distribution knowledge to override if clearly misleading.
7. Vague Localities: For vague entries like "Guinea", "West Indies", "South America", or "Africa", use knowledge of the species' known geographic range to narrow down the most likely country. If the range spans multiple countries and the locality cannot be pinned, assign "?".
8. Unresolvable: Assign "?" only when the ambiguity genuinely cannot be resolved even with distribution knowledge.

Output format:
Respond ONLY with a valid JSON array. No explanation, no markdown, no code fences — just the raw JSON array. One object per entry, in the exact same order as the input, numbered to match:
[
  {"n": 1, "corrected_country": "<country name or ?>", "comment": "<one concise sentence>"},
  {"n": 2, "corrected_country": "<country name or ?>", "comment": "<one concise sentence>"},
  ...
]"""


# ══════════════════════════════════════════════════════════════════════════════
#  CHECKPOINT SYSTEM
#
#  The checkpoint file (checkpoint.json) is the script's memory across runs.
#  It records:
#    - Which rows have been processed (by sheet and row index)
#    - The corrected country and comment for each processed row
#    - How much has been spent today
#    - What date "today" is (so the daily counter resets on a new day)
#
#  Every time a batch finishes, the checkpoint is saved immediately. If the
#  script is interrupted mid-run, you lose at most one batch of ~20 rows.
# ══════════════════════════════════════════════════════════════════════════════

def load_checkpoint():
    """
    Load progress from a previous run, or create a fresh state if none exists.
    Automatically resets the daily spend counter when a new day is detected.
    """
    if Path(CHECKPOINT_FILE).exists():
        with open(CHECKPOINT_FILE, "r") as f:
            state = json.load(f)
        # If it's a new calendar day, reset the daily spend so we get a fresh budget
        if state.get("date") != str(date.today()):
            print(f"  New day detected — resetting daily spend counter.")
            state["date"] = str(date.today())
            state["daily_spend_usd"] = 0.0
        print(f"  Loaded checkpoint from previous run.")
        return state
    else:
        print("  No checkpoint found — starting fresh.")
        # Initialize empty results dicts for all four sheets
        return {
            "date": str(date.today()),
            "daily_spend_usd": 0.0,
            "results": {
                "Reptiles":   {},
                "Mammals":    {},
                "Amphibians": {},
                "Birds":      {}
            }
        }


def save_checkpoint(state):
    """Save progress to disk. Called after every batch."""
    with open(CHECKPOINT_FILE, "w") as f:
        json.dump(state, f, indent=2)


# ══════════════════════════════════════════════════════════════════════════════
#  DATA LOADING
#
#  Reads all four sheets from the Excel file into memory as plain Python lists.
#  Only the columns relevant to the analysis are kept: species name, year,
#  current country, and locality text.
# ══════════════════════════════════════════════════════════════════════════════

def load_excel_data(filepath):
    """
    Read the Excel file and return a dict of { sheet_name: [list of row dicts] }.
    Row indices (1-based, matching Excel row numbers minus the header) are
    preserved so results can be written back to the correct cells.
    """
    print(f"\n  Reading {filepath}...")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    all_data = {}

    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Map column names to their indices (handles varying column order)
        headers = {str(h): i for i, h in enumerate(rows[0]) if h is not None}

        sheet_rows = []
        for row_idx, row in enumerate(rows[1:], start=1):  # row_idx=1 means first data row
            def get(col_name, fallback=""):
                i = headers.get(col_name)
                v = row[i] if i is not None else None
                return str(v).strip() if v is not None and str(v).strip() not in ("", "None") else fallback

            locality = get("Locality")
            # Treat these phrases as "no locality" regardless of capitalisation
            if locality.lower() in ("na", "n/a", "none"):
                locality = ""

            sheet_rows.append({
                "row_index": row_idx,
                "binomial":  get("Binomial",          "Unknown"),
                "year":      get("Description_Year",  "?"),
                "country":   get("Country",            "?"),
                "locality":  locality,
            })

        all_data[sheet_name] = sheet_rows
        print(f"    {sheet_name}: {len(sheet_rows):,} rows")

    wb.close()
    return all_data


# ══════════════════════════════════════════════════════════════════════════════
#  PRE-FILTERING (FREE STEP)
#
#  Before spending any API tokens, we automatically handle rows where no
#  useful locality text exists. There's nothing Claude can do with a blank
#  locality, so we assign "?" for free and skip these in the main loop.
#
#  This typically covers 10–20% of rows, saving real money.
# ══════════════════════════════════════════════════════════════════════════════

# Phrases that indicate no usable locality information
TRIVIAL_LOCALITY_PHRASES = {
    "", "?", "na", "n/a", "none", "unknown",
    "type locality unknown", "locality unknown", "locality not given",
    "no data", "not stated", "not given", "without locality",
    "without locality data", "no type locality", "type locality not stated",
}

def pre_filter_trivial_rows(all_data, state):
    """
    Scan all rows and auto-assign '?' to any row with no usable locality.
    This is free — no API calls needed.
    """
    auto_count = 0
    for sheet_name, rows in all_data.items():
        processed = state["results"][sheet_name]
        for row in rows:
            idx = str(row["row_index"])
            if idx not in processed:
                if row["locality"].strip().lower() in TRIVIAL_LOCALITY_PHRASES:
                    processed[idx] = {
                        "corrected_country": "?",
                        "comment": "No locality data available."
                    }
                    auto_count += 1

    if auto_count > 0:
        print(f"\n  Auto-assigned '?' to {auto_count:,} rows with no locality data (free — no API calls).")
        save_checkpoint(state)


# ══════════════════════════════════════════════════════════════════════════════
#  API CALL
#
#  Sends a batch of rows to Claude and returns structured results.
#
#  The rows are formatted into a numbered plain-text list (easy for Claude
#  to read) and Claude returns a JSON array with one result per row.
#  Position-based matching (row 1 in → result 1 out) makes parsing robust.
# ══════════════════════════════════════════════════════════════════════════════

def build_user_message(rows):
    """
    Format a list of row dicts into the text we send to Claude.
    Each entry is numbered so the output JSON can be matched by position.
    """
    lines = []
    for i, row in enumerate(rows, start=1):
        lines.append(
            f"{i}. {row['binomial']} (described {row['year']}) | "
            f"Current country: {row['country']}\n"
            f"   Locality: {row['locality']}"
        )
    return "\n".join(lines)


def extract_json_array(text):
    """
    Extract a JSON array from Claude's response even if it contains extra text.
    Claude is instructed to output only JSON, but this makes parsing robust.
    """
    # Try direct parse first
    try:
        return json.loads(text.strip())
    except json.JSONDecodeError:
        pass
    # Fall back to finding the first [...] block in the text
    match = re.search(r'\[.*\]', text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group())
        except json.JSONDecodeError:
            pass
    return None


def call_claude(client, rows, attempt=1):
    """
    Send a batch of rows to Claude and return (results_list, cost_usd).

    On a rate limit error, waits 60 seconds and retries once.
    On a JSON parse error, returns (None, 0) so the caller can skip this batch
    and retry it on the next run.
    """
    user_message = build_user_message(rows)

    try:
        response = client.messages.create(
            model=MODEL,
            max_tokens=4096,   # Generous limit — 20 rows of JSON output needs ~500-800 tokens
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_message}]
        )

        # --- Count cost ---
        input_tokens  = response.usage.input_tokens
        output_tokens = response.usage.output_tokens
        cost = (input_tokens * INPUT_PRICE_PER_TOKEN) + (output_tokens * OUTPUT_PRICE_PER_TOKEN)

        # --- Parse response ---
        raw_text = next((b.text for b in response.content if b.type == "text"), "")
        parsed   = extract_json_array(raw_text)

        if parsed is None:
            print(f"\n    Warning: Could not parse JSON from response.")
            print(f"    Raw response snippet: {raw_text[:200]!r}")
            return None, 0.0

        return parsed, cost

    except anthropic.RateLimitError:
        if attempt == 1:
            print(f"\n    Rate limit hit — waiting 60 seconds before retrying...")
            time.sleep(60)
            return call_claude(client, rows, attempt=2)
        else:
            print(f"\n    Rate limit hit twice in a row. Saving progress and stopping.")
            raise

    except anthropic.APIError as e:
        print(f"\n    API error: {e}. Skipping this batch.")
        return None, 0.0


# ══════════════════════════════════════════════════════════════════════════════
#  OUTPUT WRITING
#
#  Writes a new Excel file containing all original columns plus two new ones:
#  "Corrected_Country" and "Comment". Rows not yet processed get blank cells.
#
#  This function is called at the end of every run (and on Ctrl+C) so you
#  always have an up-to-date output file to inspect.
# ══════════════════════════════════════════════════════════════════════════════

def write_output(input_filepath, output_filepath, state):
    """
    Create a new Excel file with original data + two result columns per sheet.
    Unprocessed rows get empty cells in the new columns.
    """
    print(f"\n  Writing results to {output_filepath} ...")
    wb_in  = openpyxl.load_workbook(input_filepath)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # Remove default blank sheet

    for sheet_name in wb_in.sheetnames:
        ws_in  = wb_in[sheet_name]
        ws_out = wb_out.create_sheet(title=sheet_name)
        results = state["results"].get(sheet_name, {})
        rows_written = 0

        for row_idx, row in enumerate(ws_in.iter_rows(values_only=True)):
            if row_idx == 0:
                # Header row: append two new column names
                ws_out.append(list(row) + ["Corrected_Country", "Comment"])
            else:
                result = results.get(str(row_idx), {})
                ws_out.append(
                    list(row) + [
                        result.get("corrected_country", ""),
                        result.get("comment", "")
                    ]
                )
                rows_written += 1

    wb_out.save(output_filepath)
    wb_in.close()
    print(f"  Saved.")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN LOOP
#
#  This is where everything comes together. The loop:
#    1. Finds all rows that haven't been processed yet
#    2. Groups them into batches of ROWS_PER_CALL (default: 20)
#    3. Sends each batch to Claude
#    4. Saves results to the checkpoint after every batch
#    5. Stops when the daily spending limit is reached
#
#  On the next run, the checkpoint is loaded and processing continues from
#  exactly where it stopped, skipping all already-completed rows.
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("\n" + "═" * 62)
    print("  Tetrapod Type Locality Validator")
    print("═" * 62)

    # ── Validate API key ─────────────────────────────────────────────────────
    if not ANTHROPIC_API_KEY:
        print("\n  ERROR: No Anthropic API key found.")
        print("  Set one with:  export ANTHROPIC_API_KEY='sk-ant-...'")
        print("  Or paste it into the ANTHROPIC_API_KEY variable in this script.")
        sys.exit(1)

    # ── Graceful shutdown on Ctrl+C ──────────────────────────────────────────
    # If you press Ctrl+C mid-run, the script saves progress before quitting
    # so you don't lose work.
    shutdown_requested = {"value": False}
    def handle_ctrl_c(sig, frame):
        print("\n\n  Interrupted! Saving progress before quitting...")
        shutdown_requested["value"] = True
    signal.signal(signal.SIGINT, handle_ctrl_c)

    # ── Load checkpoint and data ─────────────────────────────────────────────
    state    = load_checkpoint()
    all_data = load_excel_data(INPUT_FILE)

    # Ensure results dict has keys for any sheet that might be missing
    for sheet_name in all_data:
        if sheet_name not in state["results"]:
            state["results"][sheet_name] = {}

    # Auto-handle rows with no locality (free, no API cost)
    pre_filter_trivial_rows(all_data, state)

    # ── Summarise current status ─────────────────────────────────────────────
    total_rows = sum(len(rows) for rows in all_data.values())
    total_done = sum(len(v) for v in state["results"].values())
    total_left = total_rows - total_done

    print(f"\n  Progress : {total_done:,} / {total_rows:,} rows complete")
    print(f"  Remaining: {total_left:,} rows")
    print(f"  Today's spend so far: ${state['daily_spend_usd']:.4f} / ${DAILY_BUDGET_USD:.2f}")

    # Rough estimate of days remaining
    if total_left > 0:
        est_cost_total = total_left * 0.0015  # ~$0.0015 per row (rough estimate)
        est_days = max(1, int((est_cost_total - (DAILY_BUDGET_USD - state['daily_spend_usd'])) / DAILY_BUDGET_USD) + 1)
        print(f"  Est. remaining cost: ~${est_cost_total:.2f} (~{est_days} more day(s) at ${DAILY_BUDGET_USD}/day)")

    if total_left == 0:
        print("\n  All rows already processed! Writing final output...")
        write_output(INPUT_FILE, OUTPUT_FILE, state)
        print("\n  Done. Check", OUTPUT_FILE)
        return

    if state["daily_spend_usd"] >= DAILY_BUDGET_USD:
        print(f"\n  Today's budget is already used up. Run again tomorrow.")
        write_output(INPUT_FILE, OUTPUT_FILE, state)
        return

    # ── Main processing loop ─────────────────────────────────────────────────
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    rows_processed_this_session = 0
    batches_this_session = 0

    print(f"\n  Starting... (Ctrl+C to stop safely at any time)\n")

    for sheet_name, rows in all_data.items():
        if shutdown_requested["value"]:
            break
        if state["daily_spend_usd"] >= DAILY_BUDGET_USD:
            break

        processed_this_sheet = state["results"][sheet_name]
        pending = [r for r in rows if str(r["row_index"]) not in processed_this_sheet]

        if not pending:
            continue

        total_batches = (len(pending) + ROWS_PER_CALL - 1) // ROWS_PER_CALL
        print(f"  ── {sheet_name} ── ({len(pending):,} rows remaining, {total_batches} batches)")

        for batch_num, batch_start in enumerate(range(0, len(pending), ROWS_PER_CALL), start=1):
            if shutdown_requested["value"]:
                break
            if state["daily_spend_usd"] >= DAILY_BUDGET_USD:
                print(f"\n  Daily budget of ${DAILY_BUDGET_USD:.2f} reached.")
                print(f"  Run the script again tomorrow to continue.")
                break

            batch = pending[batch_start : batch_start + ROWS_PER_CALL]
            pct   = (total_done + rows_processed_this_session) / total_rows * 100

            print(
                f"  [{pct:5.1f}%] {sheet_name} batch {batch_num}/{total_batches} "
                f"(rows {batch[0]['row_index']}–{batch[-1]['row_index']}) "
                f"| spent ${state['daily_spend_usd']:.4f}",
                end="", flush=True
            )

            results, cost = call_claude(client, batch)

            if results is None:
                # Parse failed — these rows will be retried on the next run
                print(f"  [skipped — will retry next run]")
                continue

            # Match results to rows by position (result[0] → batch[0], etc.)
            # If Claude returns fewer results than rows, only the matched ones are saved.
            for i, item in enumerate(results):
                if i >= len(batch):
                    break
                row = batch[i]
                processed_this_sheet[str(row["row_index"])] = {
                    "corrected_country": str(item.get("corrected_country", "?")).strip(),
                    "comment":           str(item.get("comment", "")).strip()
                }

            state["daily_spend_usd"] += cost
            rows_processed_this_session += len(batch)
            batches_this_session += 1

            print(f"  → ${cost:.4f}")

            # Save checkpoint after every batch — never lose more than one batch of work
            save_checkpoint(state)

    # ── Write output Excel ────────────────────────────────────────────────────
    write_output(INPUT_FILE, OUTPUT_FILE, state)

    # ── Final summary ─────────────────────────────────────────────────────────
    total_done_now = sum(len(v) for v in state["results"].values())
    total_left_now = total_rows - total_done_now
    print(f"\n{'═' * 62}")
    print(f"  Session complete!")
    print(f"  Rows processed this session : {rows_processed_this_session:,}")
    print(f"  Total done                  : {total_done_now:,} / {total_rows:,}")
    print(f"  Today's spend               : ${state['daily_spend_usd']:.4f}")
    if total_left_now > 0:
        print(f"\n  {total_left_now:,} rows still pending.")
        print(f"  Run the script again tomorrow to continue.")
    else:
        print(f"\n  All {total_rows:,} rows complete!")
    print(f"  Output file: {OUTPUT_FILE}")
    print("═" * 62 + "\n")


if __name__ == "__main__":
    main()
